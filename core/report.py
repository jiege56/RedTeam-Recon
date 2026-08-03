#!/usr/bin/env python3
# coding: utf-8
"""
RedTeam信息收集 报告生成器
将所有收集结果汇总为一个 Excel 文件，总览页面直接列出所有详细信息。
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[Report] 请先安装 openpyxl: pip install openpyxl")
    raise


# 样式定义
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 分类颜色
CATEGORY_COLORS = {
    "WebServer": PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
    "Language": PatternFill(start_color="FFF3E6", end_color="FFF3E6", fill_type="solid"),
    "Framework": PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
    "CMS": PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
    "Frontend": PatternFill(start_color="F3E6FF", end_color="F3E6FF", fill_type="solid"),
    "Backend": PatternFill(start_color="E6FFF3", end_color="E6FFF3", fill_type="solid"),
    "Component": PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid"),
    "Security": PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid"),
}

# 高风险颜色
RISK_HIGH_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
RISK_MED_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")


class ReportGenerator:
    """Excel 报告生成器 - 总览页面直接列出所有详细信息。"""

    def __init__(self, config):
        self.config = config

    def generate(self, results: Dict[str, Any], output_dir: Path) -> Path:
        """生成 Excel 报告，总览页面包含所有详细信息。"""
        wb = Workbook()
        wb.remove(wb.active)

        # 创建总览工作表（包含所有详细信息）
        self._create_summary_sheet(wb, results)

        # 获取目标名称，用于报告文件名
        target = results.get("target", "unknown")
        company_info = results.get("company_info", {})
        company_name = company_info.get("company_name", "") if company_info else ""

        # 生成报告文件名
        if company_name and target and company_name != target:
            # 有企业名称和目标：企业名称_目标_信息收集报告
            safe_company = self._safe_filename(company_name)
            safe_target = self._safe_filename(target)
            report_name = f"{safe_company}_{safe_target}_信息收集报告.xlsx"
        elif company_name:
            # 只有企业名称：企业名称_信息收集报告
            safe_company = self._safe_filename(company_name)
            report_name = f"{safe_company}_信息收集报告.xlsx"
        elif target:
            # 只有目标：目标_信息收集报告
            safe_target = self._safe_filename(target)
            report_name = f"{safe_target}_信息收集报告.xlsx"
        else:
            # 默认
            report_name = f"RedTeam_Report_{results.get('timestamp', 'unknown')}.xlsx"

        # 保存
        report_path = output_dir / report_name
        wb.save(str(report_path))

        # 同时保存 JSON
        json_path = output_dir / "unified_data.json"
        self._save_json(results, json_path)

        return report_path

    def _safe_filename(self, name: str) -> str:
        """生成安全的文件名，移除特殊字符。"""
        import re
        # 移除或替换不安全的字符
        safe = re.sub(r'[<>:"/\\|?*]', '_', name)
        # 移除前后空格
        safe = safe.strip()
        # 限制长度
        if len(safe) > 50:
            safe = safe[:50]
        return safe if safe else "unknown"

    def _apply_header_style(self, ws, row, col_count):
        """应用表头样式。"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    def _apply_cell_style(self, ws, row, col):
        """应用单元格样式。"""
        cell = ws.cell(row=row, column=col)
        cell.font = CELL_FONT
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER
        return cell

    def _auto_width(self, ws, min_width=12, max_width=60):
        """自动调整列宽。"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    max_len = max(max_len, cell_len)
                except:
                    pass
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

    def _write_section_header(self, ws, row, title, col_count=8):
        """写入章节标题。"""
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13, color="2F5496")
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).border = Border(bottom=Side(style="medium", color="2F5496"))
        return row + 1

    def _create_summary_sheet(self, wb: Workbook, results: Dict[str, Any]):
        """创建总览工作表 - 包含所有详细信息。"""
        ws = wb.create_sheet("RedTeam信息收集报告", 0)

        # ========== 基本信息 ==========
        row = 1
        ws.cell(row=row, column=1, value="RedTeam 信息收集报告").font = Font(bold=True, size=16, color="2F5496")
        row += 2

        ws.cell(row=row, column=1, value="目标:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get("target", ""))
        row += 1
        ws.cell(row=row, column=1, value="目标类型:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get("target_type", ""))
        row += 1
        ws.cell(row=row, column=1, value="扫描时间:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=results.get("timestamp", ""))
        row += 2

        # ========== 企业信息（如果有） ==========
        company_info = results.get("company_info")
        if company_info:
            row = self._write_section_header(ws, row, f"🏢 企业信息收集: {company_info.get('company_name', '')}")
            ws.cell(row=row, column=1, value="企业名称:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=company_info.get("company_name", ""))
            row += 1

            # FOFA查询语句
            queries = company_info.get("fofa_queries", [])
            if queries:
                ws.cell(row=row, column=1, value="FOFA查询:").font = Font(bold=True)
                ws.cell(row=row, column=2, value=" | ".join(queries))
                row += 2

            # 发现的域名
            domains = company_info.get("domains", [])
            if domains:
                ws.cell(row=row, column=1, value=f"发现域名 ({len(domains)} 个):").font = Font(bold=True, color="2F5496")
                row += 1
                headers = ["序号", "域名"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=h)
                self._apply_header_style(ws, row, 2)
                row += 1
                for idx, domain in enumerate(domains, 1):
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=domain)
                    for col in range(1, 3):
                        self._apply_cell_style(ws, row, col)
                    row += 1
                row += 1

            # 发现的IP
            ips = company_info.get("ips", [])
            if ips:
                ws.cell(row=row, column=1, value=f"发现IP ({len(ips)} 个):").font = Font(bold=True, color="2F5496")
                row += 1
                headers = ["序号", "IP地址"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=h)
                self._apply_header_style(ws, row, 2)
                row += 1
                for idx, ip in enumerate(ips, 1):
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=ip)
                    for col in range(1, 3):
                        self._apply_cell_style(ws, row, col)
                    row += 1
                row += 1

            # ICP备案信息
            icp_info = company_info.get("icp_info", [])
            if icp_info:
                ws.cell(row=row, column=1, value=f"ICP备案信息 ({len(icp_info)} 条):").font = Font(bold=True, color="2F5496")
                row += 1
                headers = ["序号", "ICP号/域名", "来源"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=h)
                self._apply_header_style(ws, row, 3)
                row += 1
                for idx, icp in enumerate(icp_info, 1):
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=icp.get("icp_number", icp.get("domain", "")))
                    ws.cell(row=row, column=3, value=icp.get("source", ""))
                    for col in range(1, 4):
                        self._apply_cell_style(ws, row, col)
                    row += 1
                row += 1

            # Whois信息
            whois_info = company_info.get("whois_info", [])
            if whois_info:
                ws.cell(row=row, column=1, value=f"Whois信息 ({len(whois_info)} 条):").font = Font(bold=True, color="2F5496")
                row += 1
                headers = ["域名", "注册商", "注册人", "邮箱", "创建时间"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=h)
                self._apply_header_style(ws, row, 5)
                row += 1
                for whois in whois_info:
                    ws.cell(row=row, column=1, value=whois.get("domain", ""))
                    ws.cell(row=row, column=2, value=whois.get("registrar", ""))
                    ws.cell(row=row, column=3, value=whois.get("registrant", ""))
                    ws.cell(row=row, column=4, value=whois.get("email", ""))
                    ws.cell(row=row, column=5, value=whois.get("creation_date", ""))
                    for col in range(1, 6):
                        self._apply_cell_style(ws, row, col)
                    row += 1
                row += 1

            # 邮箱地址
            emails = company_info.get("emails", [])
            if emails:
                ws.cell(row=row, column=1, value=f"邮箱地址 ({len(emails)} 个):").font = Font(bold=True, color="2F5496")
                row += 1
                headers = ["序号", "邮箱", "来源"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=h)
                self._apply_header_style(ws, row, 3)
                row += 1
                for idx, email in enumerate(emails, 1):
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=email.get("email", ""))
                    ws.cell(row=row, column=3, value=email.get("source", ""))
                    for col in range(1, 4):
                        self._apply_cell_style(ws, row, col)
                    row += 1
                row += 1

            # GitHub泄露
            github_leaks = company_info.get("github_leaks", [])
            if github_leaks:
                ws.cell(row=row, column=1, value=f"GitHub代码泄露 ({len(github_leaks)} 条):").font = Font(bold=True, color="FF0000")
                row += 1
                headers = ["序号", "文件", "仓库", "路径", "链接"]
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=h)
                self._apply_header_style(ws, row, 5)
                row += 1
                for idx, leak in enumerate(github_leaks, 1):
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=leak.get("file", ""))
                    ws.cell(row=row, column=3, value=leak.get("repo", ""))
                    ws.cell(row=row, column=4, value=leak.get("path", ""))
                    ws.cell(row=row, column=5, value=leak.get("url", ""))
                    # 高亮显示敏感文件
                    file_name = leak.get("file", "").lower()
                    if any(x in file_name for x in ['password', 'secret', 'key', 'token', 'config', '.env']):
                        for col in range(1, 6):
                            ws.cell(row=row, column=col).fill = RISK_HIGH_FILL
                    for col in range(1, 6):
                        self._apply_cell_style(ws, row, col)
                    row += 1
                row += 1

        # ========== 统计概览 ==========
        fp_count = sum(len(fp.get("fingerprints", [])) for fp in results.get("fingerprints", []))
        frameworks = []
        for fp_data in results.get("fingerprints", []):
            for fp in fp_data.get("fingerprints", []):
                name = fp.get("name", "")
                if name and name not in frameworks:
                    frameworks.append(name)

        row = self._write_section_header(ws, row, "📊 统计概览")
        stats = [
            ("子域名数量", len(results.get("subdomains", []))),
            ("IP地址数量", len(results.get("ips", []))),
            ("资产数量", len(results.get("assets", []))),
            ("框架指纹", fp_count),
            ("目录扫描命中", len(results.get("dir_hits", []))),
            ("漏洞/弱口令", len(results.get("brute_findings", []))),
        ]
        for key, value in stats:
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        if frameworks:
            ws.cell(row=row, column=1, value="识别框架:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=", ".join(frameworks))
            row += 1
        row += 1

        # ========== 子域名列表 ==========
        subdomains = results.get("subdomains", [])
        if subdomains:
            row = self._write_section_header(ws, row, f"🌐 子域名列表 ({len(subdomains)} 个)")
            headers = ["序号", "子域名", "IP地址", "状态码", "标题"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            for idx, sub in enumerate(subdomains, 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=sub.get("subdomain", ""))
                ws.cell(row=row, column=3, value=sub.get("ip", ""))
                ws.cell(row=row, column=4, value=sub.get("status", ""))
                ws.cell(row=row, column=5, value=sub.get("title", ""))
                for col in range(1, 6):
                    self._apply_cell_style(ws, row, col)
                row += 1
            row += 1

        # ========== IP地址列表 ==========
        ips = results.get("ips", [])
        if ips:
            row = self._write_section_header(ws, row, f"📍 IP地址列表 ({len(ips)} 个)")
            headers = ["序号", "IP地址"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            for idx, ip in enumerate(ips, 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=ip)
                for col in range(1, 3):
                    self._apply_cell_style(ws, row, col)
                row += 1
            row += 1

        # ========== 端口扫描结果 ==========
        portscan_results = results.get("portscan_results", [])
        if portscan_results:
            row = self._write_section_header(ws, row, f"🔓 端口扫描结果")
            headers = ["序号", "IP地址", "端口", "状态", "服务", "Banner"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            idx = 0
            for scan_result in portscan_results:
                target_ip = scan_result.get("ip", scan_result.get("target", ""))
                for port_info in scan_result.get("open_ports", []):
                    idx += 1
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=target_ip)
                    ws.cell(row=row, column=3, value=port_info.get("port", ""))
                    ws.cell(row=row, column=4, value=port_info.get("state", ""))
                    ws.cell(row=row, column=5, value=port_info.get("service", ""))
                    ws.cell(row=row, column=6, value=port_info.get("banner", "")[:50])

                    # 高危端口标红
                    port = port_info.get("port", 0)
                    if port in (21, 23, 445, 1433, 1521, 3306, 3389, 5432, 6379, 27017):
                        ws.cell(row=row, column=3).fill = RISK_HIGH_FILL

                    for col in range(1, 7):
                        self._apply_cell_style(ws, row, col)
                    row += 1
            row += 1

        # ========== 资产清单 ==========
        assets = results.get("assets", [])
        # 过滤掉端口扫描来源的资产（已在上面显示）
        filtered_assets = [a for a in assets if a.get("source") != "builtin_portscan"]
        if filtered_assets:
            row = self._write_section_header(ws, row, f"🖥️ 资产清单 ({len(filtered_assets)} 个)")
            headers = ["序号", "URL/IP", "端口", "状态码", "标题", "服务", "类型"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            for idx, asset in enumerate(filtered_assets, 1):
                ws.cell(row=row, column=1, value=idx)
                if asset.get("type") == "web":
                    ws.cell(row=row, column=2, value=asset.get("url", ""))
                    ws.cell(row=row, column=4, value=asset.get("status", ""))
                    ws.cell(row=row, column=5, value=asset.get("title", ""))
                    ws.cell(row=row, column=6, value=asset.get("server", ""))
                    ws.cell(row=row, column=7, value="Web")
                elif asset.get("type") == "service":
                    ws.cell(row=row, column=2, value=asset.get("ip", ""))
                    ws.cell(row=row, column=3, value=asset.get("port", ""))
                    ws.cell(row=row, column=6, value=asset.get("service", ""))
                    ws.cell(row=row, column=7, value="服务")
                else:
                    ws.cell(row=row, column=2, value=asset.get("raw", str(asset)))
                    ws.cell(row=row, column=7, value="其他")
                for col in range(1, 8):
                    self._apply_cell_style(ws, row, col)
                row += 1
            row += 1

        # ========== 框架指纹详情 ==========
        fingerprints = results.get("fingerprints", [])
        if fingerprints:
            row = self._write_section_header(ws, row, f"🔍 框架指纹识别 ({fp_count} 个)")
            headers = ["序号", "目标URL", "框架/组件", "分类", "识别证据", "状态码", "Server"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            idx = 0
            for fp_data in fingerprints:
                url = fp_data.get("url", "")
                status_code = fp_data.get("status_code", "")
                server = fp_data.get("server", "")

                for fp in fp_data.get("fingerprints", []):
                    idx += 1
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=url)
                    ws.cell(row=row, column=3, value=fp.get("name", ""))
                    ws.cell(row=row, column=4, value=fp.get("category", ""))
                    ws.cell(row=row, column=5, value=fp.get("evidence", ""))
                    ws.cell(row=row, column=6, value=status_code)
                    ws.cell(row=row, column=7, value=server)

                    # 分类颜色
                    category = fp.get("category", "")
                    if category in CATEGORY_COLORS:
                        for col in range(1, 8):
                            ws.cell(row=row, column=col).fill = CATEGORY_COLORS[category]

                    for col in range(1, 8):
                        self._apply_cell_style(ws, row, col)
                    row += 1
            row += 1

        # ========== 目录扫描结果 ==========
        dir_hits = results.get("dir_hits", [])
        if dir_hits:
            row = self._write_section_header(ws, row, f"📁 目录扫描结果 ({len(dir_hits)} 个)")
            headers = ["序号", "路径", "状态码", "大小"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            for idx, hit in enumerate(dir_hits, 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=hit.get("path", ""))
                ws.cell(row=row, column=3, value=hit.get("status", ""))
                ws.cell(row=row, column=4, value=hit.get("size", ""))

                # 高风险状态码标红
                status = hit.get("status", "")
                if status in ("200", "403"):
                    ws.cell(row=row, column=3).fill = RISK_HIGH_FILL

                for col in range(1, 5):
                    self._apply_cell_style(ws, row, col)
                row += 1
            row += 1

        # ========== 漏洞/弱口令 ==========
        brute_findings = results.get("brute_findings", [])
        if brute_findings:
            row = self._write_section_header(ws, row, f"🔓 漏洞/弱口令发现 ({len(brute_findings)} 个)")
            headers = ["序号", "类型", "用户名", "密码", "服务"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            for idx, finding in enumerate(brute_findings, 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=finding.get("type", ""))
                ws.cell(row=row, column=3, value=finding.get("user", ""))
                ws.cell(row=row, column=4, value=finding.get("password", ""))
                ws.cell(row=row, column=5, value=finding.get("service", ""))

                # 弱口令标红
                ws.cell(row=row, column=4).fill = RISK_HIGH_FILL

                for col in range(1, 6):
                    self._apply_cell_style(ws, row, col)
                row += 1
            row += 1

        # ========== 执行步骤详情 ==========
        steps = results.get("steps", {})
        if steps:
            row = self._write_section_header(ws, row, "⚙️ 执行步骤详情")
            headers = ["工具", "状态", "返回码", "输出长度", "错误信息"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            self._apply_header_style(ws, row, len(headers))
            row += 1

            for tool_id, step_result in steps.items():
                ws.cell(row=row, column=1, value=tool_id)
                ws.cell(row=row, column=2, value="✓ 成功" if step_result.get("success") else "✗ 失败")
                ws.cell(row=row, column=3, value=step_result.get("returncode", -1))
                ws.cell(row=row, column=4, value=len(step_result.get("output", "")))
                ws.cell(row=row, column=5, value=step_result.get("error", ""))

                # 成功绿色，失败红色
                if step_result.get("success"):
                    ws.cell(row=row, column=2).font = Font(color="008000")
                else:
                    ws.cell(row=row, column=2).font = Font(color="FF0000")

                for col in range(1, 6):
                    self._apply_cell_style(ws, row, col)
                row += 1

        # 自动调整列宽
        self._auto_width(ws)

    def _save_json(self, results: Dict[str, Any], json_path: Path):
        """保存 JSON 格式结果。"""
        clean_results = {}
        for key, value in results.items():
            if key == "steps":
                clean_steps = {}
                for tool_id, step in value.items():
                    clean_step = {k: v for k, v in step.items() if k != "parsed"}
                    clean_steps[tool_id] = clean_step
                clean_results[key] = clean_steps
            else:
                clean_results[key] = value

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(clean_results, f, ensure_ascii=False, indent=2)
